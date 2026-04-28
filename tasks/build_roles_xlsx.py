from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

FONT_NAME = "Arial"
title_font = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
header_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
sub_header_font = Font(name=FONT_NAME, size=11, bold=True)
body_font = Font(name=FONT_NAME, size=10)
body_bold = Font(name=FONT_NAME, size=10, bold=True)

title_fill = PatternFill("solid", start_color="1F4E78")
header_fill = PatternFill("solid", start_color="2E75B6")
admin_fill = PatternFill("solid", start_color="D9E1F2")
manager_fill = PatternFill("solid", start_color="E2EFDA")
tech_fill = PatternFill("solid", start_color="FFF2CC")
staff_fill = PatternFill("solid", start_color="FCE4D6")
yes_fill = PatternFill("solid", start_color="C6EFCE")
no_fill = PatternFill("solid", start_color="FFC7CE")
partial_fill = PatternFill("solid", start_color="FFEB9C")

thin = Side(border_style="thin", color="999999")
border = Border(top=thin, bottom=thin, left=thin, right=thin)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header_row(ws, row, cols, fill=header_fill, font=header_font):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = center
        cell.border = border


def style_body(ws, start_row, end_row, cols, align=left):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.font.name != FONT_NAME:
                cell.font = body_font
            cell.alignment = align
            cell.border = border


# ===== Sheet 1: 角色定义 =====
ws1 = wb.active
ws1.title = "角色定义"

ws1.merge_cells("A1:D1")
ws1["A1"] = "工单报修系统 - 角色定义"
ws1["A1"].font = title_font
ws1["A1"].fill = title_fill
ws1["A1"].alignment = center
ws1.row_dimensions[1].height = 28

headers1 = ["角色 ID", "英文标识", "中文名称", "角色定位"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=i, value=h)
style_header_row(ws1, 3, 4)

roles_data = [
    (1, "ADMIN", "系统管理员", "最高权限，几乎可执行所有操作；唯一可进入管理后台"),
    (2, "PROPERTY_MANAGER", "行政经理", "业务管理层，负责工单流转、人员管理、数据导出"),
    (3, "MAINTENANCE_STAFF", "维修员", "仅维修执行权限，不能提交工单；按部门接单"),
    (4, "PROPERTY_STAFF", "办美员工", "默认角色，新用户注册即为此角色；可提交工单与复核"),
]
for i, row in enumerate(roles_data, 4):
    for j, val in enumerate(row, 1):
        ws1.cell(row=i, column=j, value=val)

style_body(ws1, 4, 7, 4)
for r, fill in zip(range(4, 8), [admin_fill, manager_fill, tech_fill, staff_fill]):
    ws1.cell(row=r, column=1).fill = fill
    ws1.cell(row=r, column=1).alignment = center
    ws1.cell(row=r, column=1).font = body_bold
    ws1.cell(row=r, column=2).alignment = center
    ws1.cell(row=r, column=3).alignment = center
    ws1.cell(row=r, column=3).font = body_bold

ws1.column_dimensions["A"].width = 10
ws1.column_dimensions["B"].width = 22
ws1.column_dimensions["C"].width = 16
ws1.column_dimensions["D"].width = 60

ws1.cell(row=9, column=1, value="说明：")
ws1.cell(row=9, column=1).font = body_bold
ws1.merge_cells("A10:D10")
ws1["A10"] = ("权限分两层：① 角色硬编码权限（云函数中按 role_id 判断）；"
              "② 5 个可动态分配的模块权限：submit_work_orders / review_work_orders / "
              "view_analytics / manage_users / configure_system，由管理员在「角色与权限」页面配置。")
ws1["A10"].alignment = left
ws1["A10"].font = body_font
ws1.row_dimensions[10].height = 45

# ===== Sheet 2: 权限对照速查表 =====
ws2 = wb.create_sheet("权限对照速查表")

ws2.merge_cells("A1:E1")
ws2["A1"] = "权限对照速查表"
ws2["A1"].font = title_font
ws2["A1"].fill = title_fill
ws2["A1"].alignment = center
ws2.row_dimensions[1].height = 28

headers2 = ["功能", "系统管理员 (1)", "行政经理 (2)", "维修员 (3)", "办美员工 (4)"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 5)
ws2.cell(row=3, column=2).fill = admin_fill
ws2.cell(row=3, column=2).font = sub_header_font
ws2.cell(row=3, column=3).fill = manager_fill
ws2.cell(row=3, column=3).font = sub_header_font
ws2.cell(row=3, column=4).fill = tech_fill
ws2.cell(row=3, column=4).font = sub_header_font
ws2.cell(row=3, column=5).fill = staff_fill
ws2.cell(row=3, column=5).font = sub_header_font

matrix = [
    ("提交工单", "✅", "✅", "❌", "✅"),
    ("查看工单", "全部", "全部", "部门内", "大多数"),
    ("编辑工单", "✅", "✅（已提报）", "❌", "自己提交的（已提报）"),
    ("删除工单", "✅", "自己提交的", "❌", "自己提交的"),
    ("接单 / 维修操作", "✅", "❌", "✅（部门匹配）", "❌"),
    ("完成维修", "✅", "❌", "✅", "❌"),
    ("复核验收", "✅", "❌", "❌", "部分（自己/同岗提交）"),
    ("催接单 / 催维修", "✅", "✅", "❌", "✅"),
    ("催复核", "✅", "✅", "✅", "❌"),
    ("导出工单 Excel", "❌", "✅（≤5000 条）", "❌", "❌"),
    ("用户列表 / 创建用户", "✅", "✅", "❌", "❌"),
    ("删除 / 启用 / 停用用户", "✅", "❌", "❌", "❌"),
    ("角色权限配置", "✅", "❌", "❌", "❌"),
    ("公告管理（增删改发布）", "✅", "❌", "❌", "❌"),
    ("审计日志查看", "✅", "❌", "❌", "❌"),
    ("系统配置（增/改/批量）", "✅", "❌", "❌", "❌"),
    ("数据字典管理", "✅", "❌", "❌", "❌"),
    ("数据分析查看", "✅", "可授权", "❌", "可授权"),
    ("反馈管理", "全部", "自己的", "自己的", "自己的"),
    ("进入管理后台 (pages/admin)", "✅", "❌", "❌", "❌"),
]

for i, row in enumerate(matrix, 4):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        cell.font = body_font
        cell.border = border
        if j == 1:
            cell.alignment = left
            cell.font = body_bold
        else:
            cell.alignment = center
            if val == "✅":
                cell.fill = yes_fill
            elif val == "❌":
                cell.fill = no_fill
            else:
                cell.fill = partial_fill

ws2.column_dimensions["A"].width = 32
for col in ["B", "C", "D", "E"]:
    ws2.column_dimensions[col].width = 24

ws2.freeze_panes = "B4"

# ===== Sheet 3: 详细权限明细 =====
ws3 = wb.create_sheet("详细权限明细")

ws3.merge_cells("A1:E1")
ws3["A1"] = "详细权限明细（按功能域分组）"
ws3["A1"].font = title_font
ws3["A1"].fill = title_fill
ws3["A1"].alignment = center
ws3.row_dimensions[1].height = 28

headers3 = ["功能域", "具体功能", "允许角色", "限制条件", "代码出处"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, 5)

details = [
    ("一、工单管理", "提交工单", "ADMIN, MANAGER, STAFF", "维修员禁止提报", "workOrderManager/handlers/crud.js:41-42"),
    ("一、工单管理", "查看工单列表", "全部角色（按角色过滤）", "维修员仅看部门工单；办美员工已完成单仅看自己提交的", "workOrderManager/handlers/crud.js:204-216"),
    ("一、工单管理", "编辑工单详情", "ADMIN, MANAGER, STAFF", "仅已提报状态；STAFF 仅可编辑自己/办美员工提交的", "workOrderManager/handlers/crud.js:264-300"),
    ("一、工单管理", "删除工单", "ADMIN, MANAGER, STAFF", "仅已提报状态；非管理员只能删自己提交的", "workOrderManager/index.js:274-338"),

    ("二、工单流转", "接单（已提报→维修中）", "TECHNICIAN", "责任方部门匹配；并发上限 5 单", "workOrderManager/handlers/status.js:64-96"),
    ("二、工单流转", "完成维修（维修中→待复核）", "TECHNICIAN", "可提交配件清单", "workOrderManager/handlers/status.js:162-324"),
    ("二、工单流转", "复核验收（通过/返工）", "ADMIN / 提报人 / STAFF（限办美员工提交）", "复核通过后状态置为已完成", "workOrderManager/handlers/status.js:329-410"),
    ("二、工单流转", "状态跨步流转", "ADMIN", "管理员可跳过白名单流转", "workOrderManager/handlers/status.js:21-26"),

    ("三、催促功能", "催接单", "ADMIN, MANAGER, STAFF", "1 小时频控", "workOrderManager/handlers/notify.js:20-125"),
    ("三、催促功能", "催维修", "ADMIN, MANAGER, STAFF", "1 小时频控", "workOrderManager/handlers/notify.js:133-200"),
    ("三、催促功能", "催复核", "ADMIN, MANAGER, TECHNICIAN", "通知验收人", "workOrderManager/handlers/notify.js:200+"),

    ("四、导出", "导出工单 Excel", "MANAGER", "单次最多 5000 条", "workOrderManager/index.js:440-655"),

    ("五、用户管理", "用户列表查询", "ADMIN, MANAGER", "—", "userAuth/handlers/users.js:22-30"),
    ("五、用户管理", "创建用户", "ADMIN, MANAGER", "默认角色为办美员工(4)", "userAuth/handlers/users.js:110-191; helpers.js:158"),
    ("五、用户管理", "修改用户信息", "ADMIN, MANAGER", "管理员不可降低自己权限", "userAuth/handlers/users.js:196-276"),
    ("五、用户管理", "删除用户（软删）", "ADMIN", "不能删除自己", "userAuth/handlers/users.js:281-348"),
    ("五、用户管理", "启用用户", "ADMIN", "—", "userAuth/handlers/users.js:353-397"),
    ("五、用户管理", "停用用户", "ADMIN", "不能停用自己", "userAuth/handlers/users.js:402-454"),

    ("六、角色权限", "查询角色列表", "ADMIN", "—", "userAuth/handlers/system.js:18-31"),
    ("六、角色权限", "配置角色模块权限", "ADMIN", "可配置 5 个模块权限", "userAuth/handlers/system.js:36-76"),

    ("七、公告管理", "查看公告", "全部（其他角色仅看已发布）", "ADMIN 可见草稿/已发布/已下线全部", "userAuth/handlers/announcements.js:16-54"),
    ("七、公告管理", "创建/编辑/发布/下线/删除", "ADMIN", "—", "userAuth/handlers/announcements.js:74-265"),

    ("八、数据分析", "查看分析概览", "ADMIN 或 持有 view_analytics 权限", "可通过角色权限授予其他角色", "getAnalyticsOverview/index.js:49-52"),

    ("九、审计日志", "查询审计日志", "ADMIN", "支持按 action / user_id / 时间过滤", "userAuth/handlers/system.js:83-143"),

    ("十、系统配置", "查看系统配置", "全部角色", "—", "userAuth/handlers/system.js:176-191"),
    ("十、系统配置", "创建/更新/批量更新配置", "ADMIN", "—", "userAuth/handlers/system.js:196-336"),

    ("十一、数据字典", "查询字典", "全部角色", "—", "dictionaryManager/index.js"),
    ("十一、数据字典", "创建/编辑/删除字典", "ADMIN", "系统字典受保护，无法删除", "dictionaryManager/index.js:188-227"),

    ("十二、意见反馈", "创建反馈", "全部用户", "—", "feedbackManager/index.js"),
    ("十二、意见反馈", "查看反馈列表", "ADMIN（全部）/ 其他（仅自己）", "—", "feedbackManager/index.js:97-134"),
    ("十二、意见反馈", "查看反馈详情", "ADMIN 或 反馈提交者", "—", "feedbackManager/index.js:173-200"),
    ("十二、意见反馈", "删除反馈", "ADMIN 或 反馈提交者", "—", "feedbackManager/index.js:139-168"),

    ("十三、前端访问", "进入管理后台 pages/admin/*", "ADMIN", "包含用户/角色/配置/公告/审计/分析/反馈/字典", "miniprogram/behaviors/adminPage.js:14-27"),
]

for i, row in enumerate(details, 4):
    for j, val in enumerate(row, 1):
        ws3.cell(row=i, column=j, value=val)

end_row = 3 + len(details)
style_body(ws3, 4, end_row, 5)

# 给"功能域"列加分组配色
domain_colors = {
    "一、工单管理": "DDEBF7",
    "二、工单流转": "E2EFDA",
    "三、催促功能": "FFF2CC",
    "四、导出": "FCE4D6",
    "五、用户管理": "EDEDED",
    "六、角色权限": "DDEBF7",
    "七、公告管理": "E2EFDA",
    "八、数据分析": "FFF2CC",
    "九、审计日志": "FCE4D6",
    "十、系统配置": "EDEDED",
    "十一、数据字典": "DDEBF7",
    "十二、意见反馈": "E2EFDA",
    "十三、前端访问": "FFF2CC",
}
for r in range(4, end_row + 1):
    domain = ws3.cell(row=r, column=1).value
    fill_color = domain_colors.get(domain)
    if fill_color:
        ws3.cell(row=r, column=1).fill = PatternFill("solid", start_color=fill_color)
        ws3.cell(row=r, column=1).font = body_bold
        ws3.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws3.column_dimensions["A"].width = 16
ws3.column_dimensions["B"].width = 32
ws3.column_dimensions["C"].width = 38
ws3.column_dimensions["D"].width = 38
ws3.column_dimensions["E"].width = 50

ws3.freeze_panes = "A4"

# ===== Sheet 4: 模块权限说明 =====
ws4 = wb.create_sheet("模块权限说明")
ws4.merge_cells("A1:C1")
ws4["A1"] = "可动态分配的模块权限（MODULE_PERMISSIONS）"
ws4["A1"].font = title_font
ws4["A1"].fill = title_fill
ws4["A1"].alignment = center
ws4.row_dimensions[1].height = 28

headers4 = ["模块权限 Key", "中文含义", "默认拥有角色"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, 3)

modules = [
    ("submit_work_orders", "提交工单", "ADMIN, MANAGER, STAFF（维修员被硬编码禁止）"),
    ("review_work_orders", "复核工单", "ADMIN, STAFF（复核自己/同岗）"),
    ("view_analytics", "查看数据分析", "ADMIN（其他角色需管理员授权）"),
    ("manage_users", "用户管理", "ADMIN, MANAGER"),
    ("configure_system", "系统配置", "ADMIN"),
]
for i, row in enumerate(modules, 4):
    for j, val in enumerate(row, 1):
        ws4.cell(row=i, column=j, value=val)

style_body(ws4, 4, 8, 3)
for r in range(4, 9):
    ws4.cell(row=r, column=1).font = body_bold

ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 18
ws4.column_dimensions["C"].width = 56

ws4.cell(row=10, column=1, value="说明：")
ws4.cell(row=10, column=1).font = body_bold
ws4.merge_cells("A11:C11")
ws4["A11"] = ("管理员可在「角色与权限」页面（pages/admin/roles）为各角色勾选这 5 个模块权限。"
              "代码出处：cloudfunctions/userAuth/handlers/system.js:36-76")
ws4["A11"].alignment = left
ws4["A11"].font = body_font
ws4.row_dimensions[11].height = 40

output = "/Users/lvleo/Desktop/gongdanbaoxiu/tasks/角色权限清单.xlsx"
wb.save(output)
print(f"Saved: {output}")
